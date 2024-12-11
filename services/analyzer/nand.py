import io
import json
import re
import string
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from pytesseract import pytesseract
from PIL import Image
from openpyxl.utils import get_column_letter


class NandList:
    def __init__(self):
        workbook: Workbook = openpyxl.load_workbook("./data/nand_list.xlsx")
        self.sheet = workbook.active

    def find_info(self, model, lang):

        lang_column = None

        for cell in self.sheet[1]:
            if isinstance(cell.value, str):
                if cell.value == lang:
                    lang_column = cell.column
                    break
        if lang_column is None:
            return False

        rows = self.sheet.iter_rows(
            max_col=lang_column,
            values_only=True)
        for row in rows:
            if row[0] == model['name']:
                return row[lang_column - 1]
        return False

    def get_models(self) -> list:
        return [
            {
                'name': self.sheet[f'A{row}'].value,
                'row': row
             }
             for row in range(2, self.sheet.max_row + 1)
        ]