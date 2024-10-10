import io
import json
import re
import string
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from pytesseract import pytesseract
from PIL import Image
from openpyxl.utils import get_column_letter


class LogAnalyzer:
    def __init__(self, path, username, tesseract_path=None):
        workbook: Workbook = openpyxl.load_workbook("./data/panic_codes.xlsx")
        self.sheet = workbook.active
        self.log = self._read_log_file(path) if tesseract_path is None else self._read_photo(path, tesseract_path)
        self.log_dict = self.get_jsons(self.log) if tesseract_path is None else {}
        self.username = username
        self._images = {}

    @staticmethod
    def _read_log_file(path):
        with open(path, mode="r") as file:
            return file.read()

    @staticmethod
    def get_jsons(text) -> dict | None:
        try:
            text = "".join(text.split("\n")[1:])
            return json.loads(text)
        except:
            return None

    @staticmethod
    def _read_photo(path, tesseract_path):
        img = Image.open(path)
        pytesseract.tesseract_cmd = tesseract_path
        text = pytesseract.image_to_string(img, lang='eng')
        return text

    def read_images(self, sheet):
        sheet_images = sheet._images
        for image in sheet_images:
            row = image.anchor._from.row + 1
            col = get_column_letter(image.anchor._from.col)
            self._images[f'{col}{row}'] = image._data

    def get_image(self, cell):
        """Retrieves image data from a cell"""
        if cell not in self._images:
            raise ValueError("Cell {} doesn't contain an image".format(cell))
        else:
            image = io.BytesIO(self._images[cell]())
            return Image.open(image)

    def find_error_solutions(self, is_photo=False, full_version=False):
        results = []
        self.read_images(self.sheet)
        if not is_photo:
            model_column = None

            for cell in self.sheet[2]:
                if isinstance(cell.value, str):
                    if cell.value.lower().replace(" ", "") == \
                            self.log_dict["product"].lower().replace(" ", ""):
                        model_column = cell.column
                        break
            if model_column is None:
                return results
            rows = self.sheet.iter_rows(
                min_row=1,
                max_col=model_column,
                values_only=True)
            for index, row in enumerate(rows, start=1):
                if row[0] is not None:
                    result = {
                        "solutions": [],
                        "links": [],
                        "is_full": True
                    }
                    error_code = row[0].replace('“', '').replace('”', '')
                    if not full_version and error_code.find(" mini") != -1:
                        error_code = error_code[0:error_code.find(" mini")]
                        result['is_full'] = False
                    panic_string = "".join(self.log_dict.get("panicString", "").split("\n")[0:11])
                    if re.search(re.escape(str(error_code)), panic_string):
                        answer = row[model_column - 1]
                        solutions, links = self.filter_cell(answer)

                        result["solutions"].extend(solutions)
                        result["links"].extend(links)

                        try:
                            cell = f'{get_column_letter(model_column - 1)}{index}'
                            image = self.get_image(cell)
                            path = f'./{self.username}{cell}.png'

                            image.save(path)
                            result["image"] = path
                        except Exception as ex:
                            print(ex)

                        results.append(result)
            return results
        else:
            models = []
            for cell in self.sheet[2]:
                if (isinstance(cell.value, str) and
                        cell.value.lower().replace(" ", "") != "кодошибки"):
                    models.append({"name": cell.value.lower().replace(" ", ""),
                                   "num": cell.column})
            rows = self.sheet.iter_rows(
                min_row=2,
                values_only=True)

            for index, row in enumerate(rows, start=1):
                models_result = {}
                if row[0] is not None:
                    error_code = row[0].replace('“', '').replace('”', '')
                    have_panic = self.log.find(error_code)
                    if have_panic != -1:
                        model_result = { "solutions": [], "links": [] }
                        for model in models:
                            if row[model["num"] - 1] is not None:
                                answer = row[model["num"] - 1].lower().replace(" ", "")
                                solution, links = self.filter_cell(answer)

                                model_result["solutions"].extend(solution)
                                model_result["links"].extend(links)

                                try:
                                    cell = f'{get_column_letter(model["num"])}{index}'
                                    image = self.get_image(cell)
                                    path = f'./{self.username}{cell}.png'

                                    image.save(path)
                                    model_result["image"] = path
                                except:
                                    ...

                                models_result[model["name"]] = model_result
                    if models_result:
                        results.append(models_result)
            return results

    def get_model(self) -> list:
        header_row = self.sheet[1]
        model_row = self.sheet[2]

        for header_cell, model_cell in zip(header_row[1:], model_row[1:]):
            if isinstance(model_cell.value, str):
                if model_cell.value.lower().replace(" ", "") == \
                        self.log_dict["product"].lower().replace(" ", ""):
                    return [header_cell.value, model_cell.value]

    @staticmethod
    def filter_cell(text: str):
        solutions = []
        links = []
        for value in text.split(";"):
            if (value := value.strip()).startswith("http"):
                links.append(value)
            elif value:
                solutions.append(value)
        return solutions, links
