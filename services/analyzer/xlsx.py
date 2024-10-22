import pandas as pd
from openpyxl.reader.excel import load_workbook


def get_cities():
    data = pd.read_excel('data/cities.xlsx')
    result = {}

    for _, row in data.iterrows():
        country = row[0]
        city = row[1]

        if country in result:
            result[country].append(city)
        else:
            result[country] = [city]

    return result


def is_valid_panic_xlsx(path: str):
    try:
        workbook = load_workbook(filename=path)
        for sheet in workbook.worksheets:
            if path.endswith("panic_codes.xlsx"):
                print(sheet.max_row)
                for row in range(3, sheet.max_row + 1):
                    print(sheet[f'A{row}'].value)
                    print(sheet[f'A{row}'].value.strip())
                is_column_valid = all(
                    isinstance(sheet[f'A{row}'].value, str) and sheet[f'A{row}'].value.strip() != '' for row in
                    range(3, sheet.max_row + 1))

                is_rows_valid = all(
                    isinstance(sheet.cell(row=row, column=col).value, str) and sheet.cell(row=row, column=col).value.strip() != ''
                    for row in range(1, 3)  # строки 1 и 2
                    for col in range(2, sheet.max_column + 1)  # начиная с столбца 'B'
                )
                result = all([is_column_valid, is_rows_valid])
                if not result:
                    return False
            elif path.endswith("cities.xlsx"):
                for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    cell_value = row[0]
                    if not isinstance(cell_value, str) or not cell_value.strip():
                        return False
        else:
            return True
    except Exception as ex:
        print(ex)
        return False